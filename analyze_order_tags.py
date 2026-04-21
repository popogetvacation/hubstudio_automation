"""
订单标签分析脚本
分析数据库订单数据，根据规则添加标签并导出到Excel

筛选条件：status = 'To Ship' AND tracking_number 为空
"""
import os
from datetime import datetime
from typing import Dict, List, Set
import sys

# 添加 src 目录到路径以导入配置
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'src'))

from config import load_config

# 尝试导入所需库
try:
    import pyodbc
except ImportError:
    print("请安装 pyodbc: pip install pyodbc")
    exit(1)

try:
    from openpyxl import Workbook
except ImportError:
    print("请安装 openpyxl: pip install openpyxl")
    exit(1)


def get_db_connection():
    """获取数据库连接（从 config/settings.yaml 中读取路径）"""
    config = load_config()
    db_path = config.database.access_path
    driver = 'Microsoft Access Driver (*.mdb, *.accdb)'
    print(f"数据库地址：{db_path}")
    conn_str = f"DRIVER={{{driver}}};DBQ={db_path};charset=utf-8;"
    conn = pyodbc.connect(conn_str)
    conn.setdecoding(pyodbc.SQL_CHAR, encoding='latin1')
    conn.setdecoding(pyodbc.SQL_WCHAR, encoding='latin1')
    return conn


def get_all_orders(conn) -> List[Dict]:
    """获取所有订单数据（筛选 To Ship 且无追踪号）"""
    cursor = conn.cursor()
    # 筛选条件: status = 'To Ship' AND tracking_number 为空或空字符串
    cursor.execute("""
        SELECT order_sn, order_id, buyer_user_id, rating, shipping_address,
               total_price, currency, order_create_time, status, tracking_number
        FROM shopee_orders
        WHERE status = 'To Ship' AND (tracking_number IS NULL OR tracking_number = '')
    """)
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    return [dict(zip(columns, row)) for row in rows]


def get_order_items(conn) -> List[Dict]:
    """获取所有订单商品数据"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT order_sn, item_id, model_id, amount, item_name
        FROM shopee_order_items
    """)
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    return [dict(zip(columns, row)) for row in rows]


def get_order_buyers(conn) -> List[Dict]:
    """获取所有订单买家信息（包括聊天记录）"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT order_sn, buyer_user_id, buyer_username, rating as buyer_rating,
               country, city, conversation_id, user_message_text
        FROM shopee_order_buyer
    """)
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    return [dict(zip(columns, row)) for row in rows]


def check_low_score(rating) -> bool:
    """检查低分用户：顾客评分 > 0 且 < 3"""
    try:
        r = float(rating)
        return r > 0 and r < 3
    except:
        return False


def check_same_order_multi_items(order_sn: str, items: List[Dict]) -> bool:
    """检查同单多件：同一订单中任意一款产品购买数量 >= 2"""
    order_items = [item for item in items if item.get('order_sn') == order_sn]
    for item in order_items:
        amount = item.get('amount', 0)
        if amount and int(amount) >= 2:
            return True
    return False


def check_high_frequency_repurchase(order_sn: str, order_create_time,
                                     history_orders: List[tuple],
                                     order_items_map: Dict) -> bool:
    """
    检查高频复购：同一顾客在3小时内有多次下单记录，
    且购买了同款产品（item_id 相同）

    Args:
        order_sn: 当前订单号
        order_create_time: 当前订单创建时间
        history_orders: 历史订单列表，每项为 (order_sn, status, tracking_number, order_create_time) 元组
        order_items_map: 订单商品映射

    注：从历史订单中检查，包括所有状态的订单
    """
    if not order_create_time:
        return False

    # 获取当前订单的商品
    current_items = order_items_map.get(order_sn, [])
    current_item_ids = set(i.get('item_id') for i in current_items if i.get('item_id'))
    if not current_item_ids:
        return False

    # 遍历历史订单，检查是否同时满足：
    # 1. 时间差在3小时内
    # 2. 存在同款产品（item_id 相同）
    # 3. 不存在退货记录（status 为 Canceled 且有 tracking_number 视为发货后退货）
    for row in history_orders:
        hist_order_sn = row[0]
        hist_status = row[1]
        hist_tracking_number = row[2]
        hist_order_create_time = row[3]

        if not hist_order_create_time:
            continue

        # 计算时间差（小时）
        time_diff = abs((order_create_time - hist_order_create_time).total_seconds()) / 3600

        if time_diff <= 3:
            # 检查同款商品
            o_items = order_items_map.get(hist_order_sn, [])
            o_item_ids = set(i.get('item_id') for i in o_items if i.get('item_id'))

            if current_item_ids & o_item_ids:
                # 检查是否有退货记录：Canceled 且有 tracking_number 视为退货
                has_tracking = hist_tracking_number and str(hist_tracking_number).strip()
                is_canceled_with_tracking = (hist_status and hist_status.lower() == 'cancelled') and has_tracking

                if is_canceled_with_tracking:
                    # 退货重新下单，不计入高频复购
                    continue

                return True

    return False


def check_ph_remote_area(order, buyer_info) -> Dict:
    """
    检查菲律宾PH边远地区订单
    返回: {'is_remote': bool, 'has_chat': bool, 'reason': str}
    """
    rating = order.get('rating')
    total_price = order.get('total_price', 0)
    currency = order.get('currency', '')
    shipping_address = order.get('shipping_address', '') or ''

    # 条件1: 顾客评分 = 0
    try:
        rating_is_zero = rating is not None and float(rating) == 0
    except:
        rating_is_zero = False

    # 条件2: 收货地址位于 Mindanao 或 Visayas 地区
    address_lower = shipping_address.lower()
    is_mindanao = 'mindanao' in address_lower
    is_visayas = 'visayas' in address_lower or 'cebu' in address_lower or 'iloilo' in address_lower
    is_remote_region = is_mindanao or is_visayas

    # 条件3: 订单金额 > 6000 PHP
    is_high_value = currency == 'PHP' and total_price and float(total_price) > 6000

    # 如果不满足任一条件，不属于边远地区订单
    if not (rating_is_zero and is_remote_region and is_high_value):
        return {'is_remote': False, 'has_chat': False, 'reason': ''}

    # 检查是否有客服交流记录
    user_message_text = buyer_info.get('user_message_text', '') or ''
    has_chat = len(user_message_text.strip()) > 0

    if has_chat:
        return {'is_remote': False, 'has_chat': True, 'reason': '可通过'}
    else:
        return {'is_remote': True, 'has_chat': False, 'reason': '地址偏远'}


def get_buyer_history_orders(conn, order_sn: str, buyer_user_id: str) -> List[tuple]:
    """
    获取买家的全部历史订单（排除当前订单）

    Args:
        conn: 数据库连接
        order_sn: 当前订单号（需要排除）
        buyer_user_id: 买家用户ID

    Returns:
        历史订单列表，每项为 (order_sn, status, tracking_number, order_create_time) 元组
    """
    if not buyer_user_id:
        return []

    # Access ODBC 不支持参数化查询，使用字符串拼接
    escaped_buyer = buyer_user_id.replace("'", "''")
    escaped_sn = order_sn.replace("'", "''")

    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT order_sn, status, tracking_number, order_create_time
        FROM shopee_orders
        WHERE buyer_user_id = '{escaped_buyer}' AND order_sn <> '{escaped_sn}'
    """)

    return cursor.fetchall()


def check_suspicious_customer(history_orders: List[tuple]) -> bool:
    """
    检查可疑顾客：历史是否存在满足以下条件的订单：
    1. 订单状态不为 Completed
    2. 订单状态不为 Canceled
    3. 且无运单号

    Args:
        history_orders: 历史订单列表，每项为 (order_sn, status, tracking_number, order_create_time) 元组
    """
    # 判断条件：
    # 1. status 不为 'Completed'
    # 2. (status 为 'Canceled' 但有运单号) 或 (status 不为 'Canceled' 且无运单号)
    #    即：排除正常取消（Canceled + 无运单号）和已完成（Completed）
    for row in history_orders:
        status = row[1]
        tracking_number = row[2]
        if status:
            status_lower = status.lower()
            # 条件1: 订单状态不为 Completed 且不为 To Ship
            is_not_completed_or_to_ship = (status_lower != 'completed' and status_lower != 'to_ship')

            # 条件2: 检查是否为可疑订单
            # - 如果是 Canceled 但有运单号 -> 可疑（发货后取消）
            # - 如果不是 Canceled 且无运单号 -> 可疑（未完成且未发货）
            has_tracking = tracking_number and str(tracking_number).strip()
            is_canceled_with_tracking = (status_lower == 'cancelled') and has_tracking
            is_not_canceled_no_tracking = (status_lower != 'cancelled') and not has_tracking

            if is_not_completed_or_to_ship and (is_canceled_with_tracking or is_not_canceled_no_tracking):
                return True

    return False


# 税务关键词列表
TAX_KEYWORDS = [
    '税', '税务', '发票', 'invoice', 'receipt', 'tax',
    'ภาษี', 'การเก็บภาษี', 'ใบแจ้งหนี้'
]


def check_tax_request(order, buyer_info) -> bool:
    """
    检查税务相关请求
    检查订单备注或聊天记录中是否包含税务关键词
    """
    user_message_text = buyer_info.get('user_message_text', '') or ''

    text_lower = user_message_text.lower()
    for keyword in TAX_KEYWORDS:
        if keyword.lower() in text_lower:
            return True

    return False


def main():
    """主函数"""
    print("=" * 60)
    print("订单标签分析脚本")
    print("筛选条件: status = 'To Ship' AND tracking_number 为空")
    print("=" * 60)

    # 保存到 data 目录
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
    os.makedirs(data_dir, exist_ok=True)

    # 清除旧的 shopee_order_tags 文件
    print("\n[0/6] 清除旧的 shopee_order_tags 文件...")
    cleared_files = []
    if os.path.exists(data_dir):
        for fname in os.listdir(data_dir):
            if fname.startswith('shopee_order_tags') and fname.endswith('.xlsx'):
                fpath = os.path.join(data_dir, fname)
                try:
                    os.remove(fpath)
                    cleared_files.append(fname)
                except Exception as e:
                    print(f"     清除失败: {fname}, {e}")
    if cleared_files:
        print(f"     已清除 {len(cleared_files)} 个旧文件: {cleared_files}")
    else:
        print("     没有旧文件需要清除")

    # 生成时间戳（格式：20260413_105651）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # 连接数据库
    print("\n[1/6] 连接数据库...")
    conn = get_db_connection()
    print("     数据库连接成功")

    # 获取数据
    print("\n[2/6] 获取订单数据（To Ship + 无追踪号）...")
    orders = get_all_orders(conn)
    print(f"     共获取 {len(orders)} 条订单")

    print("\n[3/6] 获取订单商品数据...")
    order_items = get_order_items(conn)
    print(f"     共获取 {len(order_items)} 条商品记录")

    print("\n[4/6] 获取买家信息...")
    order_buyers = get_order_buyers(conn)
    print(f"     共获取 {len(order_buyers)} 条买家记录")

    # 构建查询索引
    print("\n[5/6] 分析订单标签...")

    # 构建买家信息映射
    buyer_info_map = {b.get('order_sn'): b for b in order_buyers}

    # 构建订单商品映射
    order_items_map = {}
    for item in order_items:
        sn = item.get('order_sn')
        if sn not in order_items_map:
            order_items_map[sn] = []
        order_items_map[sn].append(item)

    # 为每个订单计算标签
    results = []

    for order in orders:
        order_sn = order.get('order_sn', '')
        buyer_user_id = order.get('buyer_user_id')
        rating = order.get('rating')
        order_create_time = order.get('order_create_time')

        tags = []

        # 1. 低分用户
        if check_low_score(rating):
            tags.append('低分不发')

        # 2a. 同单多件
        if check_same_order_multi_items(order_sn, order_items):
            tags.append('同单多件')

        # 2b & 4. 获取买家历史订单（一次查询，供两个函数使用）
        buyer_history_orders = []
        if buyer_user_id:
            buyer_history_orders = get_buyer_history_orders(conn, order_sn, buyer_user_id)

        # 2b. 高频复购
        if buyer_history_orders and check_high_frequency_repurchase(order_sn, order_create_time, buyer_history_orders, order_items_map):
            tags.append('高频复购')

        # 3. 菲律宾边远地区
        buyer_info = buyer_info_map.get(order_sn, {})
        ph_check = check_ph_remote_area(order, buyer_info)
        if ph_check.get('is_remote'):
            tags.append('地址偏远')

        # 4. 可疑顾客（历史退货退款派送失败）
        if buyer_history_orders and check_suspicious_customer(buyer_history_orders):
            tags.append('历史退货退款派送失败')

        # 5. 税务相关
        if check_tax_request(order, buyer_info):
            tags.append('顾客税务要求')

        results.append({
            'order_sn': order_sn,
            'tags': tags
        })

    # 统计标签分布
    tag_counts = {}
    for r in results:
        for tag in r['tags']:
            tag_counts[tag] = tag_counts.get(tag, 0) + 1

    print("\n     标签统计:")
    for tag, count in sorted(tag_counts.items(), key=lambda x: -x[1]):
        print(f"       - {tag}: {count} 条")

    # 给没有标签的订单添加 pass 标签
    for r in results:
        if not r['tags']:
            r['tags'].append('pass')
    print(f"\n     有标签订单: {len(results)} / {len(results)}")

    # 导出Excel（支持拆分多个文件，每个最多1000条）
    print("\n[6/6] 导出Excel...")

    # 准备所有数据行
    all_data_rows = []
    for r in results:
        order_sn = r['order_sn']
        for tag in r['tags']:
            all_data_rows.append((order_sn, tag))

    # 每1000条拆分成一个文件
    max_rows_per_file = 1000
    file_count = (len(all_data_rows) + max_rows_per_file - 1) // max_rows_per_file
    generated_files = []

    for file_idx in range(file_count):
        start_idx = file_idx * max_rows_per_file
        end_idx = min(start_idx + max_rows_per_file, len(all_data_rows))
        batch_rows = all_data_rows[start_idx:end_idx]

        wb = Workbook()
        ws = wb.active
        ws.title = "SKU"

        # 写入表头
        ws['A1'] = '*订单号/包裹号(必填)'
        ws['B1'] = '*订单标记(必填)'

        # 写入数据
        for row_idx, (order_sn, tag) in enumerate(batch_rows, start=2):
            ws[f'A{row_idx}'] = order_sn
            ws[f'B{row_idx}'] = tag

        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # 保存文件（命名格式: shopee_order_tags_20260413_105651.xlsx）
        if file_count == 1:
            output_filename = f'shopee_order_tags_{timestamp}.xlsx'
        else:
            output_filename = f'shopee_order_tags_{timestamp}_{file_idx + 1}.xlsx'
        output_path = os.path.join(data_dir, output_filename)
        wb.save(output_path)
        generated_files.append(output_path)
        print(f"     已保存: {output_filename} ({len(batch_rows)} 条)")

    print(f"\n完成! 共生成 {len(generated_files)} 个文件")

    # 关闭连接
    conn.close()

    print("\n" + "=" * 60)
    print(f"共分析 {len(orders)} 个订单，全部添加标签")
    print("=" * 60)

    # 返回生成的文件列表，供 BigSeller 上传使用
    return generated_files


if __name__ == '__main__':
    main()