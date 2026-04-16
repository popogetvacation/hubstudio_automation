"""
Lazada 订单任务
获取订单并进行标签分析，输出 lazada_order_tags.xlsx
"""
import asyncio
import json
import os
import time
from typing import Dict, Any, List, Optional
from collections import defaultdict

from .task_base import BaseTask, TaskFactory
from ..browser.selenium_driver import HubStudioSeleniumDriver
from selenium.webdriver.support.ui import WebDriverWait
from ..api.lazada_api import LazadaAPI
from ..utils.logger import default_logger as logger


class LazadaOrderTask(BaseTask):
    """
    Lazada 订单任务

    功能：
    1. 自动登录（HubStudio 保存的账号密码）
    2. 使用 API 获取订单列表（toship 待发货）
    3. 批量获取买家地址信息
    4. 批量获取买家历史订单
    5. 批量获取买家聊天记录（用于税务关键词判断）
    6. 内存中执行标签分析
    7. 输出 lazada_order_tags.xlsx
    """

    task_name = "lazada_order"

    MAX_WORKERS = 10

    # 税务关键词
    TAX_KEYWORDS = [
        'tax', 'invoice', 'receipt', 'vat', 'gst',
        '税', '税务', '发票'
    ]

    # 菲律宾偏远地区关键词
    PH_REMOTE_KEYWORDS = ['mindanao', 'visayas', 'cebu', 'iloilo']

    def __init__(self, config: Dict[str, Any] = None):
        """
        初始化订单任务

        Args:
            config: 任务配置
                - max_pages: 最大获取页数 (默认 100)
                - page_size: 每页数量 (默认 20)
                - save_to_file: 是否保存结果到文件 (默认 True)
                - output_dir: 输出目录 (默认 ./data)
        """
        super().__init__(config)

        self.max_pages = self.config.get('max_pages', 100)
        self.page_size = self.config.get('page_size', 20)
        self.save_to_file = self.config.get('save_to_file', True)
        self.output_dir = self.config.get('output_dir', './data')

        self._auth_info = None

    def setup(self, driver: HubStudioSeleniumDriver, env_info: Dict[str, Any]):
        """前置操作：导航到目标页面，确保登录状态"""
        env_name = env_info.get('env_name', '')
        lazada_api = LazadaAPI(driver)
        target_url = f"https://{lazada_api.SELLER_CENTER_DOMAIN}/"

        logger.info(f"[LazadaOrder] 导航到: {target_url}")
        driver.goto(target_url)
        time.sleep(3)

        current_url = driver.get_current_url()
        logger.info(f"[LazadaOrder] 当前URL: {current_url}")

        cookies = driver.get_cookies()
        cookie_names = [c.get('name') for c in cookies]
        logger.info(f"[LazadaOrder] Cookies数量: {len(cookies)}")

        # 检查是否需要登录
        login_cookies = ['_m_h5_tk', 'asc_uid']
        has_login_cookie = all(c in cookie_names for c in login_cookies)

        if '/login' in current_url or not has_login_cookie:
            logger.warning(f"[LazadaOrder] 未登录或登录状态失效，等待登录...")
            self._wait_for_login(driver)
        else:
            logger.info(f"[LazadaOrder] 检测到已登录Cookie，会话已恢复")

        # 等待页面加载
        try:
            WebDriverWait(driver.driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            logger.info(f"[LazadaOrder] 页面已加载完成")
        except Exception as e:
            logger.warning(f"[LazadaOrder] 等待页面加载超时: {e}")

    def execute(self, driver: HubStudioSeleniumDriver, env_info: Dict[str, Any]) -> Dict[str, Any]:
        """执行订单获取和标签分析任务"""
        env_name = env_info.get('env_name', '')

        result = {
            'env_name': env_name,
            'orders': [],
            'total_count': 0,
            'tags_count': {},
            'output_file': ''
        }

        # 初始化 API
        lazada_api = LazadaAPI(driver)
        base_url = lazada_api.get_base_url(env_name)

        logger.info(f"[LazadaOrder] 目标URL: {base_url}")

        # 导航到订单列表页面，确保 session 激活
        order_list_url = f"https://{lazada_api.SELLER_CENTER_DOMAIN}/order"
        logger.info(f"[LazadaOrder] 导航到订单页面: {order_list_url}")
        driver.goto(order_list_url)
        time.sleep(3)

        # 获取认证信息
        self._auth_info = lazada_api.auth_info
        lazada_api.set_auth_info(self._auth_info)

        # 使用异步获取 topack 待发货订单列表
        logger.info(f"[LazadaOrder] 开始异步获取订单列表...")
        loop = asyncio.new_event_loop()
        try:
            all_orders = loop.run_until_complete(
                lazada_api.get_all_orders_async(tab="topack", max_pages=self.max_pages)
            )
        finally:
            loop.close()

        result['orders'] = all_orders
        result['total_count'] = len(all_orders)

        logger.info(f"[LazadaOrder] 订单列表获取完成: {len(all_orders)} 条")

        # 无订单时直接结束任务
        if not all_orders:
            logger.info(f"[LazadaOrder] 没有订单，任务结束")
            return result

        # 提取订单号建立映射
        order_number_map = {}  # order_number -> order_data
        for order in all_orders:
            order_number = order.get('orderNumber')
            if order_number:
                order_number_map[order_number] = order

        logger.info(f"[LazadaOrder] 提取到 {len(order_number_map)} 个订单")

        if not order_number_map:
            raise RuntimeError("订单列表中未提取到有效订单（无 orderNumber）")

        # 收集所有买家 ID
        buyer_ids = set()
        order_buyer_map = {}  # order_number -> buyer_id
        for order in all_orders:
            order_number = order.get('orderNumber')
            buyer_id = str(order.get('buyerId', ''))
            if buyer_id and buyer_id != 'None':
                buyer_ids.add(buyer_id)
                order_buyer_map[order_number] = buyer_id

        logger.info(f"[LazadaOrder] 找到 {len(buyer_ids)} 个唯一买家")

        # 2. 批量获取买家地址
        logger.info(f"[LazadaOrder] 开始获取买家地址...")
        loop = asyncio.new_event_loop()
        try:
            order_numbers = list(order_number_map.keys())
            buyer_addresses = loop.run_until_complete(
                lazada_api.get_buyer_address_batch(
                    order_numbers=order_numbers,
                    max_concurrent=self.MAX_WORKERS
                )
            )
        finally:
            loop.close()

        logger.info(f"[LazadaOrder] 获取到 {len(buyer_addresses)} 个买家地址")

        # 3. 批量获取买家历史订单
        buyer_histories = {}
        if buyer_ids:
            logger.info(f"[LazadaOrder] 开始获取买家历史订单...")
            loop = asyncio.new_event_loop()
            try:
                # 取第一个买家 ID 作为当前订单对应的买家（用于 API 调用）
                first_buyer_id = list(buyer_ids)[0] if buyer_ids else ""
                first_order_number = list(order_number_map.keys())[0] if order_number_map else ""

                buyer_histories = loop.run_until_complete(
                    lazada_api.get_buyer_history_batch(
                        buyer_id=first_buyer_id,
                        order_id=first_order_number,
                        buyer_ids=list(buyer_ids),
                        max_concurrent=self.MAX_WORKERS
                    )
                )
            finally:
                loop.close()

        logger.info(f"[LazadaOrder] 获取到 {len(buyer_histories)} 个买家的历史订单")

        # 4. 批量获取买家聊天记录（用于税务关键词检查）
        buyer_chats = {}
        if buyer_ids and first_buyer_id and first_order_number:
            logger.info(f"[LazadaOrder] 开始获取买家聊天记录...")
            for buyer_id in buyer_ids:
                try:
                    order_id = order_buyer_map.get(buyer_id, first_order_number)
                    # 为每个买家单独获取 sessionViewId
                    session_view_id = lazada_api.im_open_session(buyer_id, order_id)
                    if not session_view_id:
                        logger.warning(f"[LazadaOrder] 无法获取买家 {buyer_id} 的 sessionViewId")
                        continue

                    chats = lazada_api.get_chat_history(
                        session_view_id=session_view_id,
                        buyer_id=buyer_id,
                        order_id=order_id
                    )
                    if chats:
                        buyer_chats[buyer_id] = chats
                        logger.info(f"[LazadaOrder] 买家 {buyer_id} 聊天记录: {len(chats)} 条")
                except Exception as e:
                    logger.warning(f"[LazadaOrder] 获取买家 {buyer_id} 聊天记录失败: {e}")

        logger.info(f"[LazadaOrder] 获取到 {len(buyer_chats)} 个买家的聊天记录")

        # 5. 解析地址信息（用于偏远地区判断）
        parsed_addresses = {}
        for order_number, address_info in buyer_addresses.items():
            if address_info:
                parsed = lazada_api.parse_address(address_info)
                parsed_addresses[order_number] = parsed

        # 6. 标签分析
        logger.info(f"[LazadaOrder] 开始标签分析...")

        tags_result = self._analyze_tags(
            order_number_map,
            order_buyer_map,
            buyer_addresses,
            parsed_addresses,
            buyer_histories,
            buyer_chats
        )

        result['tags_result'] = tags_result
        result['tags_count'] = tags_result['tag_counts']

        # 7. 输出 Excel
        if self.save_to_file:
            output_file = self._save_to_excel(tags_result['tagged_orders'], env_name)
            result['output_file'] = output_file
            logger.info(f"[LazadaOrder] 标签已保存到: {output_file}")

        return result

    def _analyze_tags(self,
                      order_number_map: Dict[str, Dict],
                      order_buyer_map: Dict[str, str],
                      buyer_addresses: Dict[str, Dict],
                      parsed_addresses: Dict[str, Dict],
                      buyer_histories: Dict[str, List[Dict]],
                      buyer_chats: Dict[str, List[Dict]]) -> Dict:
        """
        分析订单标签

        Args:
            order_number_map: order_number -> 订单数据
            order_buyer_map: order_number -> buyer_id
            buyer_addresses: order_number -> 地址信息
            parsed_addresses: order_number -> 解析后的地址
            buyer_histories: buyer_id -> 历史订单列表
            buyer_chats: buyer_id -> 聊天记录

        Returns:
            标签分析结果
        """
        tagged_orders = []
        tag_counts = defaultdict(int)

        for order_number, order_data in order_number_map.items():
            buyer_id = order_buyer_map.get(order_number)
            tags = []
            address_info = buyer_addresses.get(order_number)
            parsed_addr = parsed_addresses.get(order_number, {})

            # 1. 同单多件
            if self._check_same_order_multi_items(order_data):
                tags.append('同单多件')

            # 2. 高频复购
            if buyer_id and buyer_histories.get(buyer_id):
                if self._check_high_frequency_repurchase(
                    buyer_id, order_data, buyer_histories[buyer_id]
                ):
                    tags.append('高频复购')

            # 3. 地址偏远
            if address_info and buyer_id:
                if self._check_remote_area(order_data, parsed_addr):
                    tags.append('地址偏远')

            # 4. 历史退货退款派送失败
            if buyer_id and buyer_histories.get(buyer_id):
                if self._check_suspicious_customer(buyer_histories[buyer_id]):
                    tags.append('历史退货退款派送失败')

            # 5. 顾客税务要求（检查聊天记录）
            if buyer_id and buyer_chats.get(buyer_id):
                if self._check_tax_requirement(buyer_chats[buyer_id]):
                    tags.append('税务要求')

            # 6. 低分不发 (rating > 0 && rating < 3)
            # if self._check_low_rating(order_data):
            #     tags.append('低分不发')

            # 7. 如果没有标签，添加 pass
            if not tags:
                tags.append('pass')

            # 记录标签
            for tag in tags:
                tag_counts[tag] += 1

            tagged_orders.append({
                'order_id': order_number,
                'order_sn': order_number,
                'tags': tags
            })

        logger.info(f"[LazadaOrder] 标签分析完成: {dict(tag_counts)}")

        return {
            'tagged_orders': tagged_orders,
            'tag_counts': dict(tag_counts)
        }

    def _check_same_order_multi_items(self, order_data: Dict) -> bool:
        """检查同单多件：同一订单中任意产品购买数量 >= 2"""
        skus = order_data.get('skus', [])
        for sku in skus:
            quantity = sku.get('quantity', 0)
            if quantity and int(quantity) >= 2:
                return True
        return False

    def _check_high_frequency_repurchase(self, buyer_id: str,
                                         order_data: Dict,
                                         history_orders: List[Dict]) -> bool:
        """
        检查高频复购：同一顾客在2小时内有多次下单记录，且购买了同款产品 (item_id 相同)
        """
        if not buyer_id or not order_data or not history_orders:
            return False

        # 从订单获取创建时间
        creation_time = order_data.get('creationTime', 0)

        # 获取当前订单的商品 item_id
        current_items = set()
        skus = order_data.get('skus', [])
        for sku in skus:
            # Lazada 使用 orderItemId 作为商品标识
            order_item_id = str(sku.get('orderItemId', ''))
            if order_item_id:
                current_items.add(order_item_id)

        if not current_items:
            return False

        # 遍历历史订单
        current_time = int(creation_time) if creation_time else 0

        for hist_order in history_orders:
            hist_time = hist_order.get('creationTime', 0)
            if not hist_time:
                continue

            # 计算时间差（毫秒转小时）
            time_diff = abs(current_time - int(hist_time)) / (1000 * 3600)

            if time_diff <= 2:
                # 检查同款商品
                hist_skus = hist_order.get('skus', [])
                hist_item_ids = set(str(sku.get('orderItemId', '')) for sku in hist_skus if sku.get('orderItemId'))

                if current_items & hist_item_ids:
                    logger.info(f"[LazadaOrder] 检测到高频复购: buyer={buyer_id}, 时间差={time_diff:.1f}小时")
                    return True

        return False

    def _check_remote_area(self, order_data: Dict, parsed_address: Dict) -> bool:
        """
        检查菲律宾偏远地区订单

        条件：
        1. rating = 0 (废除)
        2. 地址含 Mindanao/Visayas
        3. 金额 > 6000 PHP
        """
        # 获取订单金额
        total_retail_price = order_data.get('totalRetailPrice', '0')
        try:
            price_val = float(total_retail_price.replace(',', ''))
        except:
            price_val = 0

        # 条件3: 金额 > 6000 PHP
        is_high_value = price_val > 6000

        # 条件2: 地址位于 Mindanao 或 Visayas 地区
        full_address = parsed_address.get('full_address', '').lower()
        is_remote = any(keyword in full_address for keyword in self.PH_REMOTE_KEYWORDS)

        result =  is_remote and is_high_value
        if result:
            logger.info(f"[LazadaOrder] 检测到偏远地区: 地址={full_address[:50]}, 金额={price_val}")

        return result

    def _check_suspicious_customer(self, history_orders: List[Dict]) -> bool:
        """
        检查可疑顾客：历史订单状态不为 confirmed 且不为 unpaid
        """
        for order in history_orders:
            status = order.get('tabStatus', '') or order.get('orderStatus', '')

            if status:
                status_lower = status.lower()
                is_not_confirmed = status_lower != 'confirmed'
                is_not_unpaid = status_lower != 'unpaid'

                if is_not_confirmed and is_not_unpaid:
                    logger.info(f"[LazadaOrder] 检测到可疑顾客: status={status}")
                    return True

        return False

    def _check_tax_requirement(self, chat_messages: List[Dict]) -> bool:
        """检查聊天记录中是否包含税务关键词"""
        for message in chat_messages:
            content = ''

            # 优先从 body.templateData 解析（JSON 字符串）
            body = message.get('body', {})
            template_data = body.get('templateData', '')
            if template_data:
                try:
                    template_obj = json.loads(template_data)
                    content = template_obj.get('txt', '')
                except (json.JSONDecodeError, TypeError):
                    pass

            # 如果没有，尝试直接从 body 获取
            if not content:
                content = body.get('txt', '')

            # 最后尝试直接从 message 获取 content
            if not content:
                content = message.get('content', '')

            if content:
                content_lower = content.lower()
                for keyword in self.TAX_KEYWORDS:
                    if keyword.lower() in content_lower:
                        logger.info(f"[LazadaOrder] 检测到税务要求: 关键词={keyword}, 内容={content[:50]}")
                        return True
        return False

    def _check_low_rating(self, order_data: Dict) -> bool:
        """
        检查是否为低分顾客：顾客评分 > 0 且 < 3
        """
        # Lazada 可能使用不同字段存储评分
        # 可能在 buyerInfo 或 order 数据中
        rating = order_data.get('rating', -1)

        # 如果订单中没有，尝试从买家信息获取
        if rating < 0:
            buyer_info = order_data.get('buyerInfo', {})
            rating = buyer_info.get('rating', -1)

        if rating > 0 and rating < 3:
            logger.info(f"[LazadaOrder] 检测到低分顾客: rating={rating}")
            return True

        return False

    def _save_to_excel(self, tagged_orders: List[Dict], env_name: str) -> str:
        """保存到 Excel 文件"""
        from openpyxl import Workbook

        os.makedirs(self.output_dir, exist_ok=True)

        timestamp = time.strftime('%Y%m%d_%H%M%S')
        safe_env_name = env_name.replace(' ', '_').replace('/', '_')
        filename = f"lazada_order_tags_{safe_env_name}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "SKU"

        # 写入表头
        ws['A1'] = '*订单号/包裹号(必填)'
        ws['B1'] = '*订单标记(必填)'

        # 准备数据行
        all_rows = []
        for order in tagged_orders:
            order_sn = order.get('order_sn', '')
            for tag in order.get('tags', []):
                all_rows.append((order_sn, tag))

        # 写入数据
        for row_idx, (order_sn, tag) in enumerate(all_rows, start=2):
            ws[f'A{row_idx}'] = order_sn
            ws[f'B{row_idx}'] = tag

        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        wb.save(filepath)

        # 如果有环境名称，额外保存一份 lazada_order_tags.xlsx
        if env_name:
            main_filename = "lazada_order_tags.xlsx"
            main_filepath = os.path.join(self.output_dir, main_filename)
            wb.save(main_filepath)
            logger.info(f"[LazadaOrder] 同时保存到: {main_filepath}")

        return filepath

    def _wait_for_login(self, driver: HubStudioSeleniumDriver, timeout: int = 180):
        """
        等待用户手动登录或自动点击登录按钮

        Args:
            driver: HubStudioSeleniumDriver 实例
            timeout: 等待超时时间（秒）
        """
        from ..api.lazada_api import LazadaAPI

        logger.info(f"[LazadaOrder] 开始等待登录（超时 {timeout} 秒）...")

        current_url = driver.get_current_url()
        logger.info(f"[LazadaOrder] 当前URL: {current_url}")

        # 如果 URL 不包含登录，跳转到登录页面
        if '/login' not in current_url:
            login_url = f"https://{LazadaAPI.SELLER_CENTER_DOMAIN}/login"
            logger.info(f"[LazadaOrder] 跳转到登录页面: {login_url}")
            driver.goto(login_url)
            time.sleep(3)

        # 尝试自动点击登录按钮
        login_button_selectors = [
            'button[data-spm="home_next"]',
            '.login-button',
            'button.login-button',
            'button.next-btn-primary'
        ]

        for selector in login_button_selectors:
            try:
                from selenium.webdriver.common.by import By
                elements = driver.driver.find_elements(By.CSS_SELECTOR, selector)
                if elements and len(elements) > 0:
                    for elem in elements:
                        try:
                            if elem.is_displayed() and elem.is_enabled():
                                logger.info(f"[LazadaOrder] 自动点击登录按钮: {selector}")
                                elem.click()
                                time.sleep(2)
                                break
                        except Exception:
                            continue
            except Exception:
                pass

        # 每5秒检测一次是否登录成功
        check_interval = 5
        start_time = time.time()
        elapsed = 0

        while elapsed < timeout:
            current_url = driver.get_current_url()
            cookies = driver.get_cookies()
            cookie_names = [c.get('name') for c in cookies]

            # 登录成功条件：URL不包含/login 且 有关键登录Cookie
            has_login = '_m_h5_tk' in cookie_names and 'asc_uid' in cookie_names
            if '/login' not in current_url and has_login:
                logger.info(f"[LazadaOrder] 登录成功！当前URL: {current_url}")
                return

            elapsed = int(time.time() - start_time)
            remaining = timeout - elapsed
            logger.info(f"[LazadaOrder] 等待登录中... ({elapsed}/{timeout}秒, 剩余 {remaining}秒)")

            time.sleep(check_interval)

        # 超时后检查
        current_url = driver.get_current_url()
        cookies = driver.get_cookies()
        cookie_names = [c.get('name') for c in cookies]
        has_login = '_m_h5_tk' in cookie_names and 'asc_uid' in cookie_names

        if '/login' not in current_url and has_login:
            logger.info(f"[LazadaOrder] 登录成功！当前URL: {current_url}")
        else:
            logger.warning(f"[LazadaOrder] 登录超时（{timeout}秒）")

    def teardown(self, driver: HubStudioSeleniumDriver, env_info: Dict[str, Any]):
        """后置操作"""
        pass

    def on_error(self, driver: HubStudioSeleniumDriver, env_info: Dict[str, Any], error: Exception):
        """错误处理：截图保存"""
        try:
            error_dir = "./screenshots/errors"
            os.makedirs(error_dir, exist_ok=True)

            env_name = env_info.get('env_name', 'unknown')
            safe_name = env_name.replace(' ', '_').replace('/', '_')
            error_path = f"{error_dir}/{safe_name}_error.png"
            driver.screenshot(error_path)
            logger.info(f"[LazadaOrder] 错误截图已保存: {error_path}")
        except Exception:
            pass


# 注册任务
TaskFactory.register(LazadaOrderTask)
