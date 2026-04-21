"""
Tokopedia/TikTok Shop 订单任务
获取订单并进行标签分析，输出 tiktok_order_tags.xlsx
"""
import asyncio
import os
import time
from typing import Dict, Any, List, Optional
from collections import defaultdict

from .task_base import BaseTask, TaskFactory
from ..browser.selenium_driver import HubStudioSeleniumDriver
from selenium.webdriver.support.ui import WebDriverWait
from ..api.tiktok_api import TokopediaAPI
from ..utils.logger import default_logger as logger


class TokopediaOrderTask(BaseTask):
    """
    Tokopedia/TikTok Shop 订单任务

    功能：
    1. 自动登录（HubStudio 保存的账号密码）
    2. 使用 API 获取订单列表
    3. 批量获取买家联系信息（地址）
    4. 批量获取买家历史订单
    5. 内存中执行标签分析
    6. 输出 tiktok_order_tags.xlsx
    """

    task_name = "tokopedia_order"

    MAX_WORKERS = 10

    # 税务关键词
    TAX_KEYWORDS = [
        '税', '税务', '发票', 'invoice', 'receipt', 'tax',
        'ภาษี', 'การเก็บภาษี', 'ใบแจ้งหนี้'
    ]

    # 菲律宾偏远地区关键词
    PH_REMOTE_KEYWORDS = ['mindanao', 'visayas', 'cebu', 'iloilo']

    def __init__(self, config: Dict[str, Any] = None):
        """
        初始化订单任务

        Args:
            config: 任务配置
                - max_pages: 最大获取页数 (默认 100)
                - page_size: 每页数量 (默认 20)
                - save_to_file: 是否保存结果到文件 (默认 True)
                - save_to_excel: 是否生成 Excel 文件 (默认 False)
                - output_dir: 输出目录 (默认 ./data)
        """
        super().__init__(config)

        self.max_pages = self.config.get('max_pages', 100)
        self.page_size = self.config.get('page_size', 20)
        self.save_to_file = self.config.get('save_to_file', True)
        self.save_to_excel = self.config.get('save_to_excel', False)  # 新增：是否生成 Excel
        self.output_dir = self.config.get('output_dir', './data')

        self._auth_info = None

        # 追踪已检查的买家，避免重复打印日志
        self._checked_buyers = set()
        self._suspicious_checked_buyers = set()

    def setup(self, driver: HubStudioSeleniumDriver, env_info: Dict[str, Any]):
        """前置操作：导航到目标页面，确保登录状态"""
        env_name = env_info.get('env_name', '')
        tiktok_api = TokopediaAPI(driver)
        target_url = tiktok_api.get_base_url(env_name)
        driver.goto(target_url)
        time.sleep(3)

        current_url = driver.get_current_url()
        logger.info(f"[TokopediaOrder] 导航到目标页面: {target_url}")

        cookies = driver.get_cookies()
        cookie_names = [c.get('name') for c in cookies]
        logger.info(f"[TokopediaOrder] 当前URL: {current_url}")
        logger.info(f"[TokopediaOrder] Cookies数量: {len(cookies)}")

        # 检查是否需要登录
        login_cookies = ['SELLER_TOKEN', 'UNIFIED_SELLER_TOKEN']
        has_login_cookie = any(c in cookie_names for c in login_cookies)

        if '/login' in current_url or not has_login_cookie:
            logger.warning(f"[TokopediaOrder] 未登录或登录状态失效，等待登录...")
            self._wait_for_login(driver, env_name, tiktok_api)
        else:
            logger.info(f"[TokopediaOrder] 检测到已登录Cookie，会话已恢复")

        # 等待页面加载
        try:
            WebDriverWait(driver.driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            logger.info(f"[TokopediaOrder] 页面已加载完成")
        except Exception as e:
            logger.warning(f"[TokopediaOrder] 等待页面加载超时: {e}")

    def execute(self, driver: HubStudioSeleniumDriver, env_info: Dict[str, Any]) -> Dict[str, Any]:
        """执行订单获取和标签分析任务"""
        env_name = env_info.get('env_name', '')

        result = {
            'env_name': env_name,
            'orders': [],
            'total_count': 0,
            'tags_count': {},
            'output_file': ''
        }

        # 初始化 API
        tiktok_api = TokopediaAPI(driver)
        base_url = tiktok_api.get_base_url(env_name)

        logger.info(f"[TokopediaOrder] 目标URL: {base_url}")

        # 先导航到订单列表页面，确保 session 激活
        order_list_url = f"{base_url}/order"
        logger.info(f"[TokopediaOrder] 导航到订单页面: {order_list_url}")
        driver.goto(order_list_url)
        time.sleep(3)

        # 获取认证信息
        # 提取认证信息并设置
        self._auth_info = tiktok_api.auth_info
        tiktok_api.set_auth_info(self._auth_info)

        # 获取订单列表
        all_orders = tiktok_api.get_all_orders(base_url, max_pages=self.max_pages)
        result['orders'] = all_orders
        result['total_count'] = len(all_orders)

        logger.info(f"[TokopediaOrder] 订单列表获取完成: {len(all_orders)} 条")

        # 无订单时直接结束任务（API 成功返回，只是没订单）
        if not all_orders:
            logger.info(f"[TokopediaOrder] 没有订单，任务结束")
            return result

        # 2. 提取 main_order_id（订单列表 API 不返回 buyer_user_id，需要后续从买家联系信息获取）
        buyer_order_map = {}  # buyer_user_id -> [main_order_ids]
        order_buyer_map = {}  # main_order_id -> buyer_user_id
        order_data_map = {}   # main_order_id -> order_data

        for order in all_orders:
            # 从订单数据中提取 main_order_id
            main_order_id = order.get('main_order_id')
            if not main_order_id:
                # 备用：从 trade_order_module 获取
                trade_module = order.get('trade_order_module', {})
                main_order_id = trade_module.get('main_order_id')

            if not main_order_id:
                logger.warning(f"[TokopediaOrder] 订单缺少 main_order_id")
                continue

            # 注意：订单列表 API 不返回 buyer_user_id
            # 先将 main_order_id 存入，稍后从买家联系信息获取 buyer_user_id
            order_buyer_map[main_order_id] = None  # 暂时为 None
            order_data_map[main_order_id] = order

        logger.info(f"[TokopediaOrder] 提取到 {len(order_buyer_map)} 个订单")

        if not order_buyer_map:
            raise RuntimeError("订单列表中未提取到有效订单（无 main_order_id）")

        # 3. 并发获取买家联系信息（包含 buyer_user_id）
        logger.info(f"[TokopediaOrder] 开始获取买家联系信息...")

        # 3a. 获取买家联系信息（地址）
        loop = asyncio.new_event_loop()
        try:
            order_ids = list(order_buyer_map.keys())
            buyer_contacts = loop.run_until_complete(
                tiktok_api.get_buyer_contact_info_batch(
                    base_url=base_url,
                    order_ids=order_ids,
                    max_concurrent=self.MAX_WORKERS
                )
            )
        finally:
            loop.close()

        logger.info(f"[TokopediaOrder] 获取到 {len(buyer_contacts)} 个买家地址")

        # 3b. 获取买家聊天链接（包含 pigeonUid 作为 buyer_user_id）
        logger.info(f"[TokopediaOrder] 开始获取买家聊天链接...")
        loop = asyncio.new_event_loop()
        try:
            order_ids = list(order_buyer_map.keys())
            chat_links = loop.run_until_complete(
                tiktok_api.get_buyer_chat_link_batch(
                    base_url=base_url,
                    order_ids=order_ids,
                    max_concurrent=self.MAX_WORKERS
                )
            )
        finally:
            loop.close()

        logger.info(f"[TokopediaOrder] 获取到 {len(chat_links)} 个买家聊天链接")

        # 3c. 从聊天链接中提取 pigeonUid 作为 im_buyer_id，建立映射
        buyer_order_map = {}  # im_buyer_id -> [main_order_ids]
        for main_order_id, chat_info in chat_links.items():
            if chat_info:
                # pigeonUid 就是买家的 IM 用户 ID
                im_buyer_id = chat_info.get('pigeonUid')
                if im_buyer_id:
                    # 更新 order_buyer_map
                    order_buyer_map[main_order_id] = im_buyer_id
                    # 建立 im_buyer_id -> orders 映射
                    if im_buyer_id not in buyer_order_map:
                        buyer_order_map[im_buyer_id] = []
                    buyer_order_map[im_buyer_id].append(main_order_id)

        logger.info(f"[TokopediaOrder] 找到 {len(buyer_order_map)} 个唯一买家")

        # 3d. 将 im_buyer_id 转换为 oec_uid
        im_buyer_to_oec_uid_map = {}
        if buyer_order_map:
            logger.info(f"[TokopediaOrder] 开始转换 IM 买家 ID 到 OEC UID...")
            loop = asyncio.new_event_loop()
            try:
                im_buyer_ids = list(buyer_order_map.keys())
                im_buyer_to_oec_uid_map = loop.run_until_complete(
                    tiktok_api.im_buyer_ids_to_oec_uids_batch(
                        base_url=base_url,
                        im_buyer_ids=im_buyer_ids,
                        max_concurrent=self.MAX_WORKERS
                    )
                )
            finally:
                loop.close()

        logger.info(f"[TokopediaOrder] 转换完成: {len(im_buyer_to_oec_uid_map)} 个成功")

        # 3e. 使用 oec_uid 获取买家历史订单
        buyer_histories = {}  # im_buyer_id -> [orders]
        if im_buyer_to_oec_uid_map:
            logger.info(f"[TokopediaOrder] 开始获取买家历史订单...")
            loop = asyncio.new_event_loop()
            try:
                # 提取 oec_uid 并去重
                oec_uids = list(set(im_buyer_to_oec_uid_map.values()))
                # 使用 oec_uid 获取订单
                oec_uid_to_orders = loop.run_until_complete(
                    tiktok_api.get_buyer_orders_batch(
                        base_url=base_url,
                        buyer_ids=oec_uids,
                        max_concurrent=self.MAX_WORKERS
                    )
                )
                # 将结果映射回 im_buyer_id
                for im_buyer_id, oec_uid in im_buyer_to_oec_uid_map.items():
                    if oec_uid in oec_uid_to_orders:
                        buyer_histories[im_buyer_id] = oec_uid_to_orders[oec_uid]
            finally:
                loop.close()

        logger.info(f"[TokopediaOrder] 获取到 {len(buyer_histories)} 个买家的历史订单")

        # 4. 内存中标签分析
        logger.info(f"[TokopediaOrder] 开始标签分析...")

        # 预先解析地址信息（用于偏远地区判断）
        parsed_addresses = {}
        for main_order_id, contact_info in buyer_contacts.items():
            if contact_info:
                parsed = tiktok_api.parse_address(contact_info)
                parsed_addresses[main_order_id] = parsed

        tags_result = self._analyze_tags(
            order_data_map,
            order_buyer_map,
            buyer_contacts,
            buyer_histories,
            parsed_addresses
        )

        result['tags_result'] = tags_result
        result['tags_count'] = tags_result['tag_counts']

        # 始终返回 tagged_orders，供 run_scheduler 使用
        result['tagged_orders'] = tags_result['tagged_orders']

        # 5. 根据 save_to_excel 配置决定是否生成 Excel
        if self.save_to_excel:
            output_file = self._save_to_excel(tags_result['tagged_orders'], env_name)
            result['output_file'] = output_file
            logger.info(f"[TokopediaOrder] 标签已保存到: {output_file}")

        return result

    def _analyze_tags(self, order_data_map: Dict,
                      order_buyer_map: Dict[str, str],
                      buyer_contacts: Dict[str, Dict],
                      buyer_histories: Dict[str, List[Dict]],
                      parsed_addresses: Dict[str, Dict] = None) -> Dict:
        """
        分析订单标签

        Args:
            order_data_map: main_order_id -> 订单数据
            buyer_contacts: main_order_id -> 联系信息
            buyer_histories: buyer_user_id -> 历史订单列表

        Returns:
            标签分析结果
        """
        # 记录已检查过的买家，避免重复打印日志
        checked_buyers = set()
        suspicious_checked_buyers = set()

        tagged_orders = []
        tag_counts = defaultdict(int)

        for main_order_id, order_data in order_data_map.items():
            # buyer_user_id 从 order_buyer_map 获取（从买家联系信息中获取）
            buyer_user_id = order_buyer_map.get(main_order_id)

            tags = []

            # 1. 同单多件
            if self._check_same_order_multi_items(order_data):
                tags.append('同单多件')
            # 2. 高频复购 (需要买家历史订单)
            if buyer_user_id and buyer_histories.get(buyer_user_id):
                if self._check_high_frequency_repurchase(
                    buyer_user_id, main_order_id, order_data, buyer_histories[buyer_user_id]
                ):
                    tags.append('高频复购')
            # 3. 地址偏远 (需要买家联系信息)
            contact_info = buyer_contacts.get(main_order_id)
            if contact_info and buyer_user_id and parsed_addresses:
                parsed_addr = parsed_addresses.get(main_order_id, {})
                if self._check_remote_area(order_data, parsed_addr):
                    tags.append('地址偏远')

            # 4. 历史退货退款派送失败
            if buyer_user_id and buyer_histories.get(buyer_user_id):
                if self._check_suspicious_customer(buyer_user_id, main_order_id, buyer_histories[buyer_user_id]):
                    tags.append('历史退货退款派送失败')

            # 5. 顾客税务要求 (预留，目前没有聊天消息)
            # TODO: 如果有聊天消息，添加税务关键词检查

            # 不再添加 'pass' 到 tags，只通过 is_pass 标记

            # 记录标签
            for tag in tags:
                tag_counts[tag] += 1

            # 计算是否为 pass 订单（tags 为空时为 pass）
            is_pass = len(tags) == 0

            tagged_orders.append({
                'platform_order_id': main_order_id,  # 改为 platform_order_id 用于匹配
                'order_sn': order_data.get('order_sn', main_order_id),
                'tags': tags,
                'is_pass': is_pass  # 添加 is_pass 字段
            })

        return {
            'tagged_orders': tagged_orders,
            'tag_counts': dict(tag_counts)
        }

    def _check_same_order_multi_items(self, order_data: Dict) -> bool:
        """检查同单多件"""
        # items 在 sku_module 中
        sku_module = order_data.get('sku_module', [])
        for item in sku_module:
            quantity = item.get('quantity', 0)
            if quantity and int(quantity) >= 2:
                return True
        return False

    def _check_high_frequency_repurchase(self, buyer_user_id: str,
                                          current_order_id: str,
                                          order_data: Dict,
                                          history_orders: List[Dict]) -> bool:
        """
        检查高频复购：同一顾客在3小时内有多次下单记录，且购买了同款产品
        注：如果历史订单存在 reverse_module，则不作为高频复购（视为退货重新下单）
        """
        if not buyer_user_id or not order_data or not history_orders:
            return False

        # 从 trade_order_module 获取订单创建时间
        trade_module = order_data.get('trade_order_module', {})
        order_create_time = trade_module.get('create_time')  # Unix 时间戳

        # 获取当前订单的商品 SKU
        current_items = set()
        sku_module = order_data.get('sku_module', [])
        for item in sku_module:
            sku_id = item.get('sku_id')
            if sku_id:
                current_items.add(str(sku_id))

        if not current_items:
            return False

        # 遍历历史订单，检查是否同时满足：
        # 1. 不是当前订单
        # 2. 时间差在2小时内
        # 3. 存在同款产品
        # 4. 不存在 reverse_module（退货重新下单不计入高频复购）
        current_time = int(order_create_time) if order_create_time else 0

        for hist_order in history_orders:
            # 跳过当前订单
            hist_main_order_id = hist_order.get('main_order_id')
            if hist_main_order_id == current_order_id:
                continue

            hist_trade_module = hist_order.get('trade_order_module', {})
            hist_time = hist_trade_module.get('create_time')  # Unix 时间戳

            if not hist_time:
                continue

            # 计算时间差（小时）
            time_diff = abs(current_time - int(hist_time)) / 3600

            if time_diff <= 2:
                # 检查同款商品
                hist_sku_module = hist_order.get('sku_module', [])
                hist_item_ids = set(str(item.get('sku_id', '')) for item in hist_sku_module if item.get('sku_id'))

                if current_items & hist_item_ids:
                    # 检查是否有 reverse_module（退货），如果有则跳过
                    hist_reverse_module = hist_order.get('reverse_module', [])
                    if hist_reverse_module:
                        logger.info(f"[TokopediaOrder] 历史订单存在退货(reverse_module)，不标记为高频复购: "
                                  f"main_order_id={hist_main_order_id}, 时间差={time_diff:.1f}小时")
                        continue

                    logger.info(f"[TokopediaOrder] 检测到高频复购: buyer={buyer_user_id}, 时间差={time_diff:.1f}小时")
                    return True

        return False

    def _check_remote_area(self, order_data: Dict, parsed_address: Dict) -> bool:
        """
        检查菲律宾偏远地区订单

        Args:
            order_data: 订单数据
            parsed_address: 已解析的地址信息（从 parsed_addresses 传入）
        """
        # 从 price_module 获取订单金额
        price_module = order_data.get('price_module', {})
        grand_total = price_module.get('grand_total', {})
        total_price = grand_total.get('price_val', 0)
        currency = grand_total.get('currency', '')

        try:
            price_val = float(total_price)
        except:
            price_val = 0

        # 只有菲律宾(PH)的偏远地区才需要检查
        sale_region = order_data.get('trade_order_module', {}).get('sale_region', '')
        is_ph = sale_region == 'PH'
        is_high_value = is_ph and currency == 'PHP' and price_val > 6000

        # 条件2: 地址位于 Mindanao 或 Visayas 地区（使用已解析的地址）
        full_address = parsed_address.get('full_address', '').lower()

        is_remote = any(keyword in full_address for keyword in self.PH_REMOTE_KEYWORDS)

        # 只有同时满足高价和偏远地区才返回 True
        return is_high_value and is_remote

    def _check_suspicious_customer(self, buyer_id: str, current_order_id: str, history_orders: List[Dict]) -> bool:
        """
        检查可疑顾客：历史订单同时存在 reverse_module 和 logistics_info_module

        判断逻辑：只有同时存在 reverse_module 和 logistics_info_module 时才标记为可疑（确认订单已发货后退货）
        """
        # 如果该买家已检查过，直接返回 False
        if buyer_id in self._suspicious_checked_buyers:
            return False
        # 标记为已检查，避免重复
        self._suspicious_checked_buyers.add(buyer_id)
        # 去重：避免重复检测同一订单
        seen_order_ids = set()

        for order in history_orders:
            # 跳过当前订单
            hist_main_order_id = order.get('main_order_id')
            if hist_main_order_id == current_order_id:
                continue

            # 跳过已处理的订单
            if hist_main_order_id in seen_order_ids:
                continue
            seen_order_ids.add(hist_main_order_id)

            # 跳过 Awaiting shipment 状态的订单 (main_order_status = 101)
            order_status_module = order.get('order_status_module', [])
            if order_status_module:
                main_order_status = order_status_module[0].get('main_order_status')
                if main_order_status == 101:  # Awaiting shipment
                    logger.info(f"[TokopediaOrder] 跳过 Awaiting shipment 状态的订单: {hist_main_order_id}")
                    continue

            reverse_module = order.get('reverse_module', [])
            logistics_info_module = order.get('logistics_info_module', [])

            if reverse_module and logistics_info_module:
                # 同时存在 reverse_module 和 logistics_info_module，确认是顾客退货
                logger.info(f"[TokopediaOrder] 检测到可疑顾客: main_order_id={hist_main_order_id}")
                self._suspicious_checked_buyers.add(buyer_id)
                return True

        return False

    def _save_to_excel(self, tagged_orders: List[Dict], env_name: str) -> str:
        """保存到 Excel 文件"""
        from openpyxl import Workbook

        os.makedirs(self.output_dir, exist_ok=True)

        timestamp = time.strftime('%Y%m%d_%H%M%S')
        safe_env_name = env_name.replace(' ', '_').replace('/', '_')
        filename = f"tiktok_order_tags_{safe_env_name}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "SKU"

        # 写入表头
        ws['A1'] = '*订单号/包裹号(必填)'
        ws['B1'] = '*订单标记(必填)'

        # 准备数据行
        all_rows = []
        for order in tagged_orders:
            order_sn = order.get('order_sn', '')
            for tag in order.get('tags', []):
                all_rows.append((order_sn, tag))

        # 写入数据
        for row_idx, (order_sn, tag) in enumerate(all_rows, start=2):
            ws[f'A{row_idx}'] = order_sn
            ws[f'B{row_idx}'] = tag

        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        wb.save(filepath)

        return filepath

    def _wait_for_login(self, driver: HubStudioSeleniumDriver, env_name: str = None, tiktok_api: TokopediaAPI = None, timeout: int = 180):
        """
        等待用户手动登录

        逻辑：
        1. 如果检测到未登录（URL包含/login或缺少登录Cookie），跳转到邮箱登录页面
        2. 等待最多3分钟，每5秒检测一次
        3. 如果URL不再包含/login，则视为登录成功

        Args:
            driver: HubStudioSeleniumDriver 实例
            timeout: 等待超时时间（秒），默认180秒（3分钟）
        """
        logger.info(f"[TokopediaOrder] 开始等待用户登录（超时 {timeout} 秒）...")

        current_url = driver.get_current_url()
        logger.info(f"[TokopediaOrder] 当前URL: {current_url}")

        # 检查是否需要跳转到邮箱登录页面
        cookies = driver.get_cookies()
        cookie_names = [c.get('name') for c in cookies]
        login_cookies = ['SELLER_TOKEN', 'UNIFIED_SELLER_TOKEN']
        has_login_cookie = any(c in cookie_names for c in login_cookies)

        if '/login' in current_url or not has_login_cookie:
            logger.info(f"[TokopediaOrder] 检测到未登录，准备跳转到邮箱登录页面...")

            # 根据环境获取正确的登录域名
            if tiktok_api and env_name:
                domain = tiktok_api._get_domain(env_name)
            else:
                # 备用方案：从当前 URL 提取
                domain = 'seller-id.tokopedia.com'  # 默认
                for code, d in TokopediaAPI.DOMAIN_MAP.items():
                    if d in current_url:
                        domain = d
                        break

            # 跳转到邮箱登录页面
            login_url = f"https://{domain}/account/login"
            logger.info(f"[TokopediaOrder] 跳转到: {login_url}")
            driver.goto(login_url)
            time.sleep(3)

            logger.info(f"[TokopediaOrder] 请在浏览器中使用邮箱扫码登录，等待中...")

        # 每5秒检测一次是否登录成功（非/login页面即为成功）
        check_interval = 5  # 秒
        start_time = time.time()
        elapsed = 0

        while elapsed < timeout:
            current_url = driver.get_current_url()
            cookies = driver.get_cookies()
            cookie_names = [c.get('name') for c in cookies]
            has_login_cookie = any(c in cookie_names for c in login_cookies)

            # 登录成功条件：URL不包含/login 且 有登录Cookie
            if '/login' not in current_url and has_login_cookie:
                logger.info(f"[TokopediaOrder] 登录成功！当前URL: {current_url}")
                return

            elapsed = int(time.time() - start_time)
            remaining = timeout - elapsed
            logger.info(f"[TokopediaOrder] 等待登录中... ({elapsed}/{timeout}秒, 剩余 {remaining}秒)")

            time.sleep(check_interval)

        # 超时后最后一次检查
        current_url = driver.get_current_url()
        cookies = driver.get_cookies()
        cookie_names = [c.get('name') for c in cookies]
        has_login_cookie = any(c in cookie_names for c in login_cookies)

        if '/login' not in current_url and has_login_cookie:
            logger.info(f"[TokopediaOrder] 登录成功！当前URL: {current_url}")
        else:
            logger.warning(f"[TokopediaOrder] 登录超时（{timeout}秒）")

    def _do_auto_login(self, driver: HubStudioSeleniumDriver):
        """
        执行自动登录操作
        1. 点击邮箱登录切换按钮
        2. 点击登录按钮
        """
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys

        logger.info(f"[TokopediaOrder] 开始自动登录...")

        # 等待页面完全加载
        time.sleep(3)

        try:
            # 1. 查找并点击邮箱登录切换按钮 (id: TikTok_Ads_SSO_Login_Email_Panel_Button)
            email_button = driver.wait_for_element_visible(
                "TikTok_Ads_SSO_Login_Email_Panel_Button", by="id", timeout=10
            )
            if email_button:
                email_button.click()
                logger.info(f"[TokopediaOrder] 已点击邮箱登录切换按钮")
                time.sleep(2)
            else:
                logger.warning(f"[TokopediaOrder] 未找到邮箱登录切换按钮，尝试其他方式...")
                # 尝试查找其他可能的切换按钮
                try:
                    alt_buttons = driver.driver.find_elements(By.CSS_SELECTOR, "[id*='email'], [class*='email'], .email-tab")
                    for btn in alt_buttons:
                        if btn.is_displayed():
                            btn.click()
                            logger.info(f"[TokopediaOrder] 点击了备选邮箱登录按钮")
                            time.sleep(2)
                            break
                except Exception as e:
                    logger.warning(f"[TokopediaOrder] 备选按钮查找失败: {e}")

            # 2. 点击登录按钮 - 尝试多种可能的选择器
            login_clicked = False

            # 尝试常见的登录按钮选择器
            login_selectors = [
                "//button[contains(@class, 'login')]",
                "//button[contains(text(), '登录')]",
                "//button[contains(text(), 'Log in')]",
                "//button[contains(text(), 'Masuk')]",
                "//button[@type='submit']",
                "[class*='login-btn']",
                "[class*='submit-btn']",
            ]

            for selector in login_selectors:
                try:
                    login_btn = driver.wait_for_element_visible(selector, by="xpath", timeout=3)
                    if login_btn and login_btn.is_displayed():
                        login_btn.click()
                        logger.info(f"[TokopediaOrder] 已点击登录按钮: {selector}")
                        login_clicked = True
                        time.sleep(3)
                        break
                except Exception:
                    continue

            if not login_clicked:
                # 尝试发送 Enter 键触发登录
                logger.info(f"[TokopediaOrder] 尝试按 Enter 键登录...")
                from selenium.webdriver.common.action_chains import ActionChains
                actions = ActionChains(driver.driver)
                actions.send_keys(Keys.RETURN).perform()
                time.sleep(3)

        except Exception as e:
            logger.error(f"[TokopediaOrder] 自动登录过程出错: {e}")

        # 等待登录完成
        time.sleep(5)
        logger.info(f"[TokopediaOrder] 自动登录操作完成，等待验证...")

    def teardown(self, driver: HubStudioSeleniumDriver, env_info: Dict[str, Any]):
        """后置操作"""
        pass

    def on_error(self, driver: HubStudioSeleniumDriver, env_info: Dict[str, Any], error: Exception):
        """错误处理：截图保存"""
        try:
            error_dir = "./screenshots/errors"
            os.makedirs(error_dir, exist_ok=True)

            env_name = env_info.get('env_name', 'unknown')
            safe_name = env_name.replace(' ', '_').replace('/', '_')
            error_path = f"{error_dir}/{safe_name}_error.png"
            driver.screenshot(error_path)
            logger.info(f"[TokopediaOrder] 错误截图已保存: {error_path}")
        except Exception:
            pass


# 注册任务
TaskFactory.register(TokopediaOrderTask)